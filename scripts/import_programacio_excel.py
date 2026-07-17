import json
import re
from datetime import date, timedelta
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\Manel Puigcerver\Downloads\programacio (1).xlsx")
OUT_SQL = Path(r"c:\MINERVA\MINERVA_HUB\minerva-strategic-ai-hub\tmp_cal_import.sql")
OUT_JSON = Path(r"c:\MINERVA\MINERVA_HUB\minerva-strategic-ai-hub\tmp_cal_import.json")

YEAR = 2026
MESES = {
    "gener": 1,
    "febrero": 2,
    "febrer": 2,
    "març": 3,
    "marzo": 3,
    "abril": 4,
    "maig": 5,
    "mayo": 5,
    "juny": 6,
    "junio": 6,
    "juliol": 7,
    "julio": 7,
    "agost": 8,
    "agosto": 8,
    "setembre": 9,
    "septiembre": 9,
    "octubre": 10,
    "novembre": 11,
    "noviembre": 11,
    "desembre": 12,
    "diciembre": 12,
}

OT_RE = re.compile(r"^(\d{4,6})[_\s\-]*(.*)$")


def parse_week(label: str) -> date | None:
    s = label.lower().replace("'", " ")
    mes = None
    for k, v in MESES.items():
        if k in s:
            mes = v
            break
    nums = [int(x) for x in re.findall(r"\d+", s)]
    if not mes or not nums:
        return None
    return date(YEAR, mes, nums[0])


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["planificador"]
    entries: list[dict] = []
    week_start: date | None = None
    orden_by_day: dict[str, int] = {}

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and "setmana" in a.lower():
            week_start = parse_week(a)
            print("WEEK", a, "->", week_start)
            # No continue: en la misma fila suelen ir OTs (col B–G).
        if week_start is None:
            continue
        for c in range(2, 8):
            v = ws.cell(r, c).value
            if v is None or r == 4:
                continue
            s = str(v).strip()
            m = OT_RE.match(s)
            if not m:
                continue
            ot = m.group(1)
            d = week_start + timedelta(days=c - 2)
            ymd = d.isoformat()
            orden = orden_by_day.get(ymd, 0)
            orden_by_day[ymd] = orden + 1
            entries.append({"fecha": ymd, "ot_numero": ot, "orden": orden})

    seen: set[tuple[str, str]] = set()
    uniq: list[dict] = []
    for e in entries:
        key = (e["fecha"], e["ot_numero"])
        if key in seen:
            continue
        seen.add(key)
        uniq.append(e)

    print("TOTAL", len(uniq))
    for e in uniq:
        print(e["fecha"], e["ot_numero"], e["orden"])

    vals = ",".join(
        f"('{e['fecha']}','{e['ot_numero']}',{e['orden']})" for e in uniq
    )
    sql = f"""
insert into public.prod_calendario_produccion_ot (fecha, ot_numero, orden)
values {vals}
on conflict (fecha, ot_numero) do update
  set orden = excluded.orden,
      updated_at = timezone('utc', now());
"""
    OUT_SQL.write_text(sql, encoding="utf-8")
    OUT_JSON.write_text(json.dumps(uniq, indent=2), encoding="utf-8")
    print("WROTE", OUT_SQL, OUT_JSON)


if __name__ == "__main__":
    main()
